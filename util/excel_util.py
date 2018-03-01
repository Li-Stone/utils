from openpyxl import *

"""
Excel 读写工具类
"""


def read_excel_to_list(file_path, sheet, start_row, end_row, name_or_index='index'):
    """
    读取excel文件中指定内容到数组中
    :param file_path:  excel文件路径
    :param sheet:  要去读的工作表的顺序
    :param start_row:  文件读取起止行
    :param end_row:  文件读取结束行
    :param name_or_index 按工作表的名字还是顺序进行读取
    :return:  指定内容的数组
    """
    try:
        wb = load_workbook(file_path)
        if name_or_index == 'name':
            sheet_name = sheet
        else:
            # 获取要读取的工作表的名字
            sheet_name = wb.get_sheet_names()[sheet - 1]
        # 根据名字获取工作表对象
        sheet = wb.get_sheet_by_name(sheet_name)
        rows = sheet.iter_rows(min_row=start_row, max_row=end_row)
        items = list()
        for row in rows:
            item = list()
            for column in row:
                if column.value is None:
                    value = ""
                else:
                    value = str(column.value)
                item.append(value)
            items.append(item)
        return items
    except Exception as e:
        print(e)


def write_list_to_excel(file_path, items, title):
    """
    将数组中内容写入到指定文件的工作表中
    :param file_path: excel文件路径
    :param items: 要写入的内容数组
    :param title sheet名字
    :return:  是否写入成功  TRUE|FALSE
    """
    try:
        wb = Workbook()
        sheet = wb.active
        sheet.title = title

        index = 0
        for item in items:
            try:
                sheet.append(item)
            except Exception as e:
                print(e)
                print(index, items[index])
            index += 1
        wb.save(file_path)
    except Exception as e:
        print(e)
        print('写入到excel失败')
        return False
    else:
        print('写入到excel成功')
        return True
